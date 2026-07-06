"""Stream the personal-property xlsx dumps: collect stats + recent rows.

Output: data/processed/pp_rows.pkl (rows with Assess Year >= 2023)
"""
import os
import time
from collections import Counter

import openpyxl
import pandas as pd

BASE = r"D:\Claude Code Projects\Building_Map\data\raw"
OUT = r"D:\Claude Code Projects\Building_Map\data\processed"

KEEP = ["PPAN", "PPAN Status", "PPAN Type", "Assess Year", "Address1", "Address2",
        "City", "Zip", "Assessed Value", "Make", "Model", "Description",
        "Prop Year", "Prop Status", "Prop Type", "Vehicle Type", "Qty"]

stats = {k: Counter() for k in ["Assess Year", "PPAN Status", "PPAN Type",
                                "Vehicle Type", "Prop Status", "Prop Type"]}
rows = []
t0 = time.time()
total = 0
for f in ["PP_Dump1.xlsx", "PP_Dump2.xlsx"]:
    wb = openpyxl.load_workbook(os.path.join(BASE, f), read_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    hdr = [str(h).strip() if h else "" for h in next(it)]
    idx = {k: hdr.index(k) for k in KEEP}
    ncol = len(hdr)
    for r in it:
        total += 1
        if len(r) < ncol:  # read_only mode truncates trailing empty cells
            r = r + (None,) * (ncol - len(r))
        if total % 500000 == 0:
            print(f"  {total} rows, kept {len(rows)}, {time.time() - t0:.0f}s", flush=True)
        try:
            ay = int(r[idx["Assess Year"]] or 0)
        except (ValueError, TypeError):
            ay = 0
        for k in stats:
            v = r[idx[k]]
            stats[k][str(v)[:20] if v is not None else ""] += 1
        if ay >= 2023:
            rows.append(tuple(r[idx[k]] for k in KEEP))
    wb.close()
    print(f"{f} done: cumulative {total} rows, kept {len(rows)}, {time.time() - t0:.0f}s", flush=True)

df = pd.DataFrame(rows, columns=KEEP)
os.makedirs(OUT, exist_ok=True)
df.to_pickle(os.path.join(OUT, "pp_rows.pkl"))
print(f"\nwrote pp_rows.pkl: {len(df)} rows of {total}")
for k, c in stats.items():
    top = ", ".join(f"{v or '<blank>'}:{n}" for v, n in c.most_common(8))
    print(f"--- {k}: {top}")
